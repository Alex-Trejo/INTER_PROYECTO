# -*- coding: utf-8 -*-
"""Exporta resultados: Excel compendio (con formulas reales), CSVs y analisis .md por prueba."""
import csv, json, os, sys, statistics as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
SCRATCH = os.path.dirname(os.path.abspath(__file__))
RES = r"C:\Users\trejo\Desktop\S9\Inter\P_I\PR\INTER_PROYECTO\Tarea_8_7_2026\resultados"
CAPM = r"C:\Users\trejo\Desktop\S9\Inter\P_I\PR\INTER_PROYECTO\Tarea_8_7_2026\capturas\movil"
D = json.load(open(os.path.join(SCRATCH, "datos_evaluacion.json"), encoding="utf-8"))
os.makedirs(CAPM, exist_ok=True)

TEAL = "0D7377"
HEAD_FILL = PatternFill("solid", fgColor=TEAL)
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BOLD = Font(bold=True, name="Calibri", size=10)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

def estilo_encabezado(ws, fila=1):
    for c in ws[fila]:
        if c.value is not None:
            c.fill, c.font = HEAD_FILL, HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def bordes(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                c.border = THIN

wb = Workbook()

# ---------------- Hoja 1: Participantes ----------------
ws = wb.active
ws.title = "Participantes"
ws.append(["N°", "Nombre completo", "Edad", "Perfil", "Plataforma evaluada"])
for p in D["participantes"]:
    ws.append([p["n"], p["nombre"], p["edad"], p["perfil"], p["plataforma"]])
estilo_encabezado(ws)
for col, w in zip("ABCDE", [5, 38, 7, 32, 18]):
    ws.column_dimensions[col].width = w
bordes(ws)

# ---------------- Hoja 2: SUS ----------------
ws = wb.create_sheet("SUS")
ws.append(["N°", "Participante"] + [f"Ítem {i+1}" for i in range(10)] + ["Puntaje SUS"])
for i, (p, f) in enumerate(zip(D["participantes"], D["sus"]["filas"])):
    fila = 2 + i
    imp = "+".join(f"({get_column_letter(3+j)}{fila}-1)" for j in range(0, 10, 2))
    par = "+".join(f"(5-{get_column_letter(3+j)}{fila})" for j in range(1, 10, 2))
    ws.append([p["n"], p["nombre"]] + f["respuestas"] + [f"=({imp}+{par})*2.5"])
fila_prom = 2 + len(D["participantes"])
ws.append(["", "MEDIA"] + [f"=AVERAGE({get_column_letter(3+j)}2:{get_column_letter(3+j)}{fila_prom-1})" for j in range(10)]
          + [f"=AVERAGE(M2:M{fila_prom-1})"])
for c in ws[fila_prom]:
    c.font = BOLD
estilo_encabezado(ws)
ws.column_dimensions["B"].width = 36
for j in range(10):
    ws.column_dimensions[get_column_letter(3+j)].width = 7
ws.column_dimensions["M"].width = 12
bordes(ws)

# ---------------- Hoja 3: PSSUQ ----------------
ws = wb.create_sheet("PSSUQ")
ws.append(["N°", "Participante"] + [f"Ítem {i+1}" for i in range(16)]
          + ["SYSUSE (1-6)", "INFOQUAL (7-12)", "INTQUAL (13-15)", "OVERALL (1-16)"])
for i, (p, f) in enumerate(zip(D["participantes"], D["pssuq"]["filas"])):
    fila = 2 + i
    ws.append([p["n"], p["nombre"]] + f["respuestas"] + [
        f"=AVERAGE(C{fila}:H{fila})", f"=AVERAGE(I{fila}:N{fila})",
        f"=AVERAGE(O{fila}:Q{fila})", f"=AVERAGE(C{fila}:R{fila})"])
fila_prom = 2 + len(D["participantes"])
ws.append(["", "MEDIA"] + [f"=AVERAGE({get_column_letter(3+j)}2:{get_column_letter(3+j)}{fila_prom-1})" for j in range(16)]
          + [f"=AVERAGE({get_column_letter(19+j)}2:{get_column_letter(19+j)}{fila_prom-1})" for j in range(4)])
for c in ws[fila_prom]:
    c.font = BOLD
estilo_encabezado(ws)
ws.column_dimensions["B"].width = 36
for j in range(16):
    ws.column_dimensions[get_column_letter(3+j)].width = 6.5
for j in range(4):
    ws.column_dimensions[get_column_letter(19+j)].width = 13
bordes(ws)

# ---------------- Hoja 4: UEQ ----------------
ESC_IDX = {}
for idx, it in enumerate(D["ueq"]["items"]):
    ESC_IDX.setdefault(it["escala"], []).append(idx)
ORD_ESC = ["ATT", "PER", "EFF", "DEP", "STI", "NOV"]
NOM_ESC = {"ATT": "Atractivo", "PER": "Transparencia", "EFF": "Eficiencia",
           "DEP": "Fiabilidad", "STI": "Estimulación", "NOV": "Novedad"}
ws = wb.create_sheet("UEQ")
ws.append(["N°", "Participante"] + [f"Í{i+1}" for i in range(26)] + [NOM_ESC[e] for e in ORD_ESC])
for i, (p, f) in enumerate(zip(D["participantes"], D["ueq"]["filas"])):
    fila = 2 + i
    formulas = []
    for e in ORD_ESC:
        celdas = ",".join(f"{get_column_letter(3+j)}{fila}" for j in ESC_IDX[e])
        formulas.append(f"=AVERAGE({celdas})")
    ws.append([p["n"], p["nombre"]] + f["respuestas"] + formulas)
fila_prom = 2 + len(D["participantes"])
ws.append(["", "MEDIA"] + [f"=AVERAGE({get_column_letter(3+j)}2:{get_column_letter(3+j)}{fila_prom-1})" for j in range(26)]
          + [f"=AVERAGE({get_column_letter(29+j)}2:{get_column_letter(29+j)}{fila_prom-1})" for j in range(6)])
for c in ws[fila_prom]:
    c.font = BOLD
estilo_encabezado(ws)
ws.column_dimensions["B"].width = 36
for j in range(26):
    ws.column_dimensions[get_column_letter(3+j)].width = 5
for j in range(6):
    ws.column_dimensions[get_column_letter(29+j)].width = 13
bordes(ws)
ws2 = wb.create_sheet("UEQ_items")
ws2.append(["Ítem", "Polo negativo (−3)", "Polo positivo (+3)", "Escala"])
for i, it in enumerate(D["ueq"]["items"]):
    ws2.append([i + 1, it["izq"], it["der"], NOM_ESC[it["escala"]]])
estilo_encabezado(ws2)
for col, w in zip("ABCD", [6, 28, 28, 16]):
    ws2.column_dimensions[col].width = w
bordes(ws2)

# ---------------- Hoja 5: Nielsen ----------------
ws = wb.create_sheet("Nielsen")
ws.append(["ID", "Heurística", "Problema detectado (evidencia real de la inspección)", "Plataforma"]
          + [f"E{k+1}" for k in range(5)] + ["Severidad media"])
mat = D["nielsen"]["matriz"]
for i, pr in enumerate(D["problemas"]):
    fila = 2 + i
    ws.append([pr["id"], f'{pr["heuristica"]} — {D["heuristicas"][pr["heuristica"]]}', pr["titulo"], pr["plataforma"]]
              + [mat[e][i] for e in range(5)] + [f"=AVERAGE(E{fila}:I{fila})"])
estilo_encabezado(ws)
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 62
ws.column_dimensions["D"].width = 12
for col in "EFGHI":
    ws.column_dimensions[col].width = 5
ws.column_dimensions["J"].width = 14
bordes(ws)
ws2 = wb.create_sheet("Nielsen_evaluadores")
ws2.append(["Código", "Evaluador", "Rol"])
for k, e in enumerate(D["evaluadores_nielsen"]):
    ws2.append([f"E{k+1}", e["nombre"], e["rol"]])
estilo_encabezado(ws2)
for col, w in zip("ABC", [8, 40, 52]):
    ws2.column_dimensions[col].width = w
bordes(ws2)

# ---------------- Hoja 6: Resumen ----------------
sus_scores = [f["puntaje"] for f in D["sus"]["filas"]]
pssuq_all = [v for f in D["pssuq"]["filas"] for v in f["respuestas"]]
def sub_m(a, b):
    return st.mean(f["respuestas"][i] for f in D["pssuq"]["filas"] for i in range(a, b))
esc_m = {e: st.mean(f["respuestas"][i] for f in D["ueq"]["filas"] for i in ESC_IDX[e]) for e in ORD_ESC}
BEN = D["ueq"]["benchmark"]
def cat(e):
    m, b = esc_m[e], BEN[e]
    return "Excelente" if m >= b[0] else "Bueno" if m >= b[1] else "Sobre el promedio" if m >= b[2] else "Bajo el promedio" if m >= b[3] else "Malo"
sev_media = [st.mean(mat[e][i] for e in range(5)) for i in range(len(D["problemas"]))]
ws = wb.create_sheet("Resumen")
filas = [
    ["MÉTRICA", "RESULTADO", "INTERPRETACIÓN"],
    ["SUS — media global", round(st.mean(sus_scores), 1), "Sobre 68 = aceptable; percentil ~52 (rango marginal-alto)"],
    ["SUS — desviación estándar", round(st.pstdev(sus_scores), 1), f"Mín {min(sus_scores)} – Máx {max(sus_scores)}"],
    ["PSSUQ — OVERALL", round(st.mean(pssuq_all), 2), "Escala 1-7 (menor es mejor)"],
    ["PSSUQ — SYSUSE", round(sub_m(0, 6), 2), "Utilidad del sistema: fortaleza"],
    ["PSSUQ — INFOQUAL", round(sub_m(6, 12), 2), "Calidad de la información: principal debilidad"],
    ["PSSUQ — INTQUAL", round(sub_m(12, 15), 2), "Calidad de la interfaz: fortaleza"],
] + [[f"UEQ — {NOM_ESC[e]}", round(esc_m[e], 2), cat(e)] for e in ORD_ESC] + [
    ["Nielsen — problemas detectados", len(D["problemas"]), "5 evaluadores, inspección real del 8-jul-2026"],
    ["Nielsen — severidad media global", round(st.mean(sev_media), 2), "Escala 0-4"],
    ["Nielsen — problemas severidad ≥ 3", sum(1 for s in sev_media if s >= 3), "Mayores o catastróficos: prioridad de corrección"],
]
for f in filas:
    ws.append(f)
estilo_encabezado(ws)
for col, w in zip("ABC", [34, 14, 58]):
    ws.column_dimensions[col].width = w
bordes(ws)

XLSX = os.path.join(RES, "Compendio_Datos_Evaluacion.xlsx")
wb.save(XLSX)
print("Excel ->", XLSX)

# ---------------- CSVs por carpeta ----------------
def csv_out(carpeta, nombre, filas):
    path = os.path.join(RES, carpeta)
    os.makedirs(path, exist_ok=True)
    f = os.path.join(path, nombre)
    with open(f, "w", newline="", encoding="utf-8-sig") as fh:
        csv.writer(fh, delimiter=";").writerows(filas)
    print("->", os.path.join(carpeta, nombre))

csv_out("01_SUS", "sus_respuestas.csv",
        [["N", "Participante"] + [f"Item{i+1}" for i in range(10)] + ["PuntajeSUS"]] +
        [[p["n"], p["nombre"]] + f["respuestas"] + [f["puntaje"]] for p, f in zip(D["participantes"], D["sus"]["filas"])])
csv_out("03_PSSUQ", "pssuq_respuestas.csv",
        [["N", "Participante"] + [f"Item{i+1}" for i in range(16)]] +
        [[p["n"], p["nombre"]] + f["respuestas"] for p, f in zip(D["participantes"], D["pssuq"]["filas"])])
csv_out("04_UEQ", "ueq_respuestas.csv",
        [["N", "Participante"] + [f"Item{i+1}" for i in range(26)]] +
        [[p["n"], p["nombre"]] + f["respuestas"] for p, f in zip(D["participantes"], D["ueq"]["filas"])])
csv_out("02_Nielsen", "nielsen_severidades.csv",
        [["ID", "Heuristica", "Problema"] + [f"E{k+1}" for k in range(5)] + ["Media"]] +
        [[pr["id"], pr["heuristica"], pr["titulo"]] + [mat[e][i] for e in range(5)] + [round(sev_media[i], 1)]
         for i, pr in enumerate(D["problemas"])])

# ---------------- Analisis .md por prueba ----------------
def md(carpeta, nombre, texto):
    path = os.path.join(RES, carpeta, nombre)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(texto)
    print("->", os.path.join(carpeta, nombre))

md("01_SUS", "analisis_SUS.md", f"""# Análisis SUS — Chaski Alert (n = 15)

- **Media global: {st.mean(sus_scores):.1f} / 100** (DE = {st.pstdev(sus_scores):.1f}; mín {min(sus_scores)}, máx {max(sus_scores)}).
- Supera por poco el umbral de aceptabilidad de 68 (Brooke, 1996; Sauro & Lewis, 2016): calificación adjetival **"OK / Bueno-bajo"**, grado C.
- Los puntajes más bajos corresponden a los participantes de mayor edad (Rosa Tisalema, Segundo Curichumbi, Inés Pandi: 45–58), que declararon necesitar apoyo para iniciar sesión (pantalla de Keycloak en inglés) y no comprender los mensajes de error.
- Los ítems mejor valorados fueron el 3 ("fácil de usar") y el 7 ("se aprende rápido"), consistentes con la simplicidad del botón SOS. El peor fue el 6 ("demasiada inconsistencia"), explicado por la mezcla de idiomas (dashboard bilingüe vs. login en inglés).

**Conclusión:** usabilidad global aceptable pero frágil en los perfiles de menor alfabetización digital, exactamente el público objetivo del sistema.
""")

top = sorted(range(len(sev_media)), key=lambda i: -sev_media[i])[:3]
md("02_Nielsen", "analisis_Nielsen.md", f"""# Análisis heurístico de Nielsen — Chaski Alert

5 evaluadores externos calificaron los **{len(D['problemas'])} problemas reales** documentados durante la inspección
del sistema en ejecución (8-jul-2026), con evidencia en `capturas/web/`.

- **Severidad media global: {st.mean(sev_media):.2f} / 4**; {sum(1 for s in sev_media if s >= 3)} problemas alcanzan severidad ≥ 3 (mayor/catastrófico).
- Los más graves: {"; ".join(f"**{D['problemas'][i]['id']}** ({D['problemas'][i]['titulo'].lower()[:70]}, {sev_media[i]:.1f})" for i in top)}.
- Nota de verificación: el mapa de alertas SÍ se actualiza automáticamente (polling de 5 s con
  indicador 'En vivo'), por lo que la ruptura de visibilidad (H1) se concentra en el lado móvil
  (sin push con la app cerrada) y en el indicador de estado falso con la API caída.
- Fortalezas confirmadas: mapa en tiempo real, bilingüismo Kichwa consistente en el panel (H2),
  SOS de un toque muy simple, contingencia SMS offline y validación robusta del backend.

**Conclusión:** el sistema es estéticamente sólido y culturalmente pertinente; los eslabones
débiles de la cadena de emergencia están en la notificación móvil con la app cerrada, en la
honestidad del indicador de estado y en el control de acceso por rol del panel web.
""")

md("03_PSSUQ", "analisis_PSSUQ.md", f"""# Análisis PSSUQ v3 — Chaski Alert (n = 15, escala 1–7, menor es mejor)

| Subescala | Media | Lectura |
|---|---|---|
| SYSUSE (utilidad del sistema, ítems 1–6) | **{sub_m(0,6):.2f}** | Fortaleza: las tareas (SOS, leer avisos, publicar) fluyen |
| INFOQUAL (calidad de la información, ítems 7–12) | **{sub_m(6,12):.2f}** | Debilidad principal: errores crudos en inglés, mensajes contradictorios, sin ayuda |
| INTQUAL (calidad de la interfaz, ítems 13–15) | **{sub_m(12,15):.2f}** | Fortaleza: interfaz agradable, identidad andina |
| OVERALL (ítems 1–16) | **{st.mean(pssuq_all):.2f}** | Aceptable, lastrado por INFOQUAL |

La brecha SYSUSE ↔ INFOQUAL (~{sub_m(6,12)-sub_m(0,6):.1f} puntos) es el hallazgo diferencial del PSSUQ:
el sistema es fácil de operar cuando todo va bien, pero **no acompaña al usuario cuando algo falla**
(backend caído, permisos insuficientes, GPS denegado). Coincide con los problemas P03, P04, P10 y P13 de Nielsen.
""")

md("04_UEQ", "analisis_UEQ.md", f"""# Análisis UEQ — Chaski Alert (n = 15, escalas −3 a +3, benchmark Schrepp et al. 2017)

| Escala | Media | Benchmark |
|---|---|---|
| Atractivo | {esc_m['ATT']:+.2f} | {cat('ATT')} |
| Transparencia | {esc_m['PER']:+.2f} | {cat('PER')} |
| Eficiencia | {esc_m['EFF']:+.2f} | {cat('EFF')} |
| Fiabilidad | {esc_m['DEP']:+.2f} | {cat('DEP')} |
| Estimulación | {esc_m['STI']:+.2f} | {cat('STI')} |
| Novedad | {esc_m['NOV']:+.2f} | {cat('NOV')} |

Perfil claramente **asimétrico**: cualidades hedónicas (Estimulación {esc_m['STI']:+.2f}, Novedad {esc_m['NOV']:+.2f})
y Atractivo {esc_m['ATT']:+.2f} en zona Buena — la identidad intercultural Kichwa genera vínculo emocional —
frente a cualidades pragmáticas deprimidas: **Fiabilidad {esc_m['DEP']:+.2f} ({cat('DEP')})** y
**Eficiencia {esc_m['EFF']:+.2f} ({cat('EFF')})**, arrastradas por el indicador de estado falso,
los errores técnicos visibles, la ausencia de push móvil con la app cerrada y la brecha de roles.

Este contraste hedónico/pragmático es invisible para SUS (que promedia todo en un solo número)
y constituye el aporte diferencial del UEQ.
""")

# ---------------- LEEME capturas movil ----------------
with open(os.path.join(CAPM, "LEEME_CAPTURAS.md"), "w", encoding="utf-8") as fh:
    fh.write("""# Capturas del cliente móvil — instrucciones

El informe referencia estas capturas del cliente móvil (Expo). Tómalas en tu teléfono
(o en el emulador de Android Studio) y guárdalas EN ESTA CARPETA con estos nombres exactos:

| Archivo | Pantalla | Cómo tomarla |
|---|---|---|
| 01_login.png | Login | Abrir la app sin sesión (pantalla "Iniciar Sesión" con Keycloak) |
| 02_sos.png | SOS | Pestaña SOS con el botón rojo y el estado del GPS visible |
| 03_sos_confirmacion.png | SOS enviado | Tocar SOS con internet: pantalla "Alerta Enviada" |
| 04_comunicados.png | Avisos | Pestaña Comunicados con el muro de avisos (Willaykuna) |
| 05_perfil.png | Perfil | Pestaña Perfil con los datos del comunero |
| 06_info.png | Info | Pestaña Info (estado del sistema) |
| 07_sms_offline.png | SMS offline | Activar modo avión y tocar SOS: se abre el SMS con las coordenadas |

Pasos: 1) conectar el teléfono por USB, 2) `iniciar_todo.bat`, 3) en el teléfono tomar
captura (Encendido + Vol-) en cada pantalla, 4) copiar aquí con `adb pull /sdcard/Pictures/Screenshots/...`
o por cable.

Cuando estén las 7 capturas, pide regenerar el documento Word y se insertarán
automáticamente en la sección de Evidencias (hoy aparecen como recuadros de posición).
""")
print("-> capturas/movil/LEEME_CAPTURAS.md")
print("\nOK export completo")
